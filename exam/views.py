from django.shortcuts import render
from .models import Ris, Smrp
from django.http import JsonResponse, HttpResponse

modaliti_choices = (
    ('CT','CT-Scan'),
    ('MR','MRI'),
    ('XR','General X-Ray'),
    ('MG','Mammography'),
    ('US','Ultrasound'),
    ('FL','Fluoroscopy'),
    ('XA','Angiography'),
    ('MF','Mobile Fluoro/OT/ERCP'),
    ('HC','Hardcopy'),
    ('DI','Digitize'),
    ('RE','Rad Resources'),
    ('RC','Ref Consult'),
)
# Create your views here.
def ris_list(request):
    ris = Ris.objects.all()
    context = {
        'ris' : ris,
        'tajuk': 'Senarai RIS',
        'modaliti': modaliti_choices
    }
    if request.htmx:
        modaliti = request.GET['modaliti']
        print(modaliti)
        context['ris'] = Ris.objects.filter(modaliti=modaliti)
        return render(request, 'ris-child.html', context)
    return render(request, 'ris.html', context )


def ris_edit(request,pk=None):
    ris = Ris.objects.get(id=pk)
    data = {
        'ris': ris
    }
    if request.method=="POST":
        smrp = request.POST['smrp']
        ris.smrp_id=smrp
        ris.save()
        data['ris'] = ris
        return render(request,'ris-single.html',data)
    
    
    return render(request,'ris-form.html', data)

def smrp_api(request):
    try:
        term = request.GET['search']
        smrp = Smrp.objects.filter(name__icontains=term)
    except:
        term = None
        smrp = Smrp.objects.all()

    result = []
    
    for a in smrp:
        result.append({'id': a.id, 'text': a.name})
    
    return JsonResponse({'results': result}, safe=False)



def get_percent(modaliti):
    data = {}
    obj = Ris.objects.filter(modaliti=modaliti)
    data['total'] = obj.count()
    taksiap = obj.filter(smrp=None).count()
    data['siap'] = obj.count() - taksiap
    return data

def progress(request):
    progress = []
    for k,v in modaliti_choices:
        progress.append({v: get_percent(k)})
    
    return JsonResponse(progress, safe=False)


from openpyxl import Workbook
from django.db.models import Case, CharField, Value, When



def export_xls(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="ris_to_smrp.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = 'RIS-SMRP'


    row_num = 0


    columns = ['Modality','Exam Code', 'Exam','SMRP' ]

    for col_num in range(0,len(columns)):
        ws.cell(row=1, column=col_num+1, value=columns[col_num])

    # Sheet body, remaining rows

    choices = dict(Ris._meta.get_field('modaliti').flatchoices)
    whens = [When(modaliti=k, then=Value(v)) for k, v in choices.items()]
    rows = (
        list(Ris.objects
        .annotate(get_modaliti_display=Case(*whens, output_field=CharField()))
        .values_list('get_modaliti_display','code', 'name','smrp__name'))
    )
    row_num=0
    for row in rows:
        row_num += 1
        for col_num in range(0,len(row)):
            ws.cell(row=row_num+1,column=col_num+1, value=str(row[col_num]))

    wb.save(response)
    return response
